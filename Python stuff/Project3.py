import mysql.connector
import cv2
import argparse
import ffmpy
import subprocess
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from pathlib import Path


def extract_timecode(video_file):
    ff = ffmpy.FFprobe(global_options="-v error -show_entries format=duration -of default=noprint_wrappers=1:nokey=1".split(), inputs={video_file: None})
    video_duration_str = ff.run(stdout=subprocess.PIPE, stderr=subprocess.PIPE)[0].strip()
    video_duration = int(float(video_duration_str) * 60)

    return video_duration


def timecode_to_frames(timecode, fps=60):
    hours, minutes, seconds, frames = map(int, timecode.split(':'))
    total_frames = frames + (seconds + minutes*60 + hours*3600) * fps
    return total_frames

def frames_to_timecode(frames, fps=60):
    seconds = frames // fps
    frames %= fps
    minutes = seconds // 60
    seconds %= 60
    hours = minutes // 60
    minutes %= 60
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}:{frames:02d}"


def create_thumbnail(video_file, idx, timecode, output_dir, fps):
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    frame_number = timecode_to_frames(timecode, fps)

    cap = cv2.VideoCapture(video_file)
    cap.set(cv2.CAP_PROP_POS_FRAMES, frame_number)

    ret, frame = cap.read()
    if ret:
        resized_frame = cv2.resize(frame, (96, 74), interpolation=cv2.INTER_AREA)
        thumbnail_path = output_dir / f"thumbnail_{idx}.png"
        cv2.imwrite(str(thumbnail_path), resized_frame)
        return thumbnail_path

    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--process', required=True, help='The file path of the video file to process')
    parser.add_argument('--output', required=True, help='The file path of the output Excel file')
    args = parser.parse_args()

    mydb = mysql.connector.connect(
        host="localhost",
        user="sqluser",
        password="93021",
        database="mydatabase"
    )

    mycursor = mydb.cursor()

    mycursor.execute("SELECT folder, frame_range FROM frames")

    wb = Workbook()
    ws = wb.active
    ws.append(['Name', 'Date', 'Folder', 'Frame Range', 'Thumbnail'])

    frame_ranges = []

    for row in mycursor.fetchall():
        folder, frame_range = row

        # Skip if the frame range is not valid
        if '-' not in frame_range:
            continue

        start_frame, end_frame = frame_range.split('-')
        if not start_frame.isdigit() or not end_frame.isdigit():
            continue

        start_frame, end_frame = int(start_frame), int(end_frame)

        # If the frame range only contains one frame number, use that frame number
        if start_frame == end_frame:
            frame_number = start_frame
        else:
            # Otherwise, use the middle frame number
            frame_number = start_frame + (end_frame - start_frame) // 2

        timecode = frames_to_timecode(frame_number, fps=60)
        thumbnail_path = create_thumbnail(args.process, ws.max_row - 1, timecode, Path(args.output).parent, fps=60)
        if thumbnail_path:
            folder = folder.replace(frame_range, f"{start_frame}-{end_frame}")
            img = Image(str(thumbnail_path))
            img.width, img.height = 96, 74
            ws.append([folder, '', frame_range, timecode])
            ws.row_dimensions[ws.max_row].height = img.height
            img.anchor = f"D{ws.max_row}"
            ws.add_image(img)

    wb.save(args.output)


if __name__ == '__main__':
    main()
